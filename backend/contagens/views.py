import json
from io import BytesIO
from decimal import Decimal

from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Count, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, render
from django.utils import timezone
from django.utils.dateparse import parse_datetime
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import Contagem, ItemContagem, NotaRecebimento


def health(_request):
    return JsonResponse({"ok": True})


@login_required
def dashboard(request):
    contagens = Contagem.objects.prefetch_related("itens")[:50]
    resumo = Contagem.objects.aggregate(
        total=Count("id"),
        total_itens=Sum("total_itens"),
        total_contado=Sum("total_contado"),
        total_recebido=Sum("total_recebido"),
    )
    por_fornecedor = (
        ItemContagem.objects.values("fornecedor")
        .annotate(
            itens=Count("id"),
            estoque=Sum("estoque_contado"),
            recebimento=Sum("recebimento_total"),
        )
        .order_by("fornecedor")
    )
    return render(
        request,
        "contagens/dashboard.html",
        {
            "contagens": contagens,
            "resumo": resumo,
            "por_fornecedor": por_fornecedor,
        },
    )


@login_required
def detalhe_contagem(request, pk):
    contagem = get_object_or_404(Contagem.objects.prefetch_related("itens__notas"), pk=pk)
    return render(request, "contagens/detalhe.html", {"contagem": contagem})


@login_required
def baixar_contagem_excel(request, pk):
    contagem = get_object_or_404(Contagem.objects.prefetch_related("itens__notas"), pk=pk)
    output = _gerar_excel_contagem(contagem)
    filename = f"contagem_{contagem.data_inicio:%Y-%m-%d}_{contagem.id}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@csrf_exempt
@require_http_methods(["GET", "POST"])
def api_contagens(request):
    if request.method == "GET":
        if not request.user.is_authenticated:
            return _unauthorized()
        data = [_contagem_json(c) for c in Contagem.objects.all()[:100]]
        return JsonResponse({"contagens": data})

    if not _token_valido(request):
        return _unauthorized()
    try:
        payload = json.loads(request.body.decode("utf-8"))
        contagem = _salvar_payload(payload)
    except (json.JSONDecodeError, KeyError, TypeError, ValueError) as exc:
        return JsonResponse({"ok": False, "erro": str(exc)}, status=400)
    return JsonResponse({"ok": True, "contagem": _contagem_json(contagem)}, status=201)


@require_http_methods(["GET"])
def api_contagem_detalhe(request, pk):
    if not request.user.is_authenticated and not _token_valido(request):
        return _unauthorized()
    contagem = get_object_or_404(Contagem.objects.prefetch_related("itens__notas"), pk=pk)
    data = _contagem_json(contagem)
    data["itens"] = [_item_json(item) for item in contagem.itens.all()]
    return JsonResponse(data)


@require_http_methods(["GET"])
def api_public_contagens(_request):
    data = [_contagem_publica_json(c) for c in Contagem.objects.prefetch_related("itens")[:100]]
    response = JsonResponse({"contagens": data})
    response["Access-Control-Allow-Origin"] = "*"
    response["Cache-Control"] = "no-store"
    return response


@transaction.atomic
def _salvar_payload(payload):
    sessao = payload["sessao"]
    itens = payload.get("itens", [])
    sessao_app_id = str(sessao["id"])

    contagem, _created = Contagem.objects.update_or_create(
        sessao_app_id=sessao_app_id,
        defaults={
            "operador_nome": sessao.get("operadorNome", ""),
            "operador_matricula": sessao.get("operadorMatricula", ""),
            "data_inicio": _parse_dt(sessao.get("dataInicio")),
            "data_fim_real": _parse_dt(sessao.get("dataFimReal")),
            "status": sessao.get("status", "recebida"),
            "versao_cadastro": sessao.get("versaoCadastro", ""),
            "aparelho": sessao.get("aparelho") or "",
            "total_itens": len(itens),
            "total_contado": _sum_decimal(itens, "estoqueContado"),
            "total_recebido": _sum_decimal(itens, "recebimentoTotal"),
            "payload_original": payload,
        },
    )

    ItemContagem.objects.filter(contagem=contagem).delete()
    for item in itens:
        material = item.get("material", {})
        item_obj = ItemContagem.objects.create(
            contagem=contagem,
            material_codigo=item.get("materialCodigo", ""),
            fornecedor=material.get("fornecedor", ""),
            material_descricao=material.get("descricao", ""),
            estoque_anterior=_decimal(item.get("estoqueAnterior")),
            estoque_contado=_decimal(item.get("estoqueContado")),
            recebimento_total=_decimal(item.get("recebimentoTotal")),
            soma_nf=_decimal(item.get("somaNotas")),
            status=item.get("status", ""),
            justificativa=item.get("justificativa") or "",
            observacao=item.get("observacao") or "",
            foto_path=item.get("fotoPath") or item.get("justificativaFotoPath") or "",
            timestamp_item=_parse_dt(item.get("timestamp")),
            payload_original=item,
        )
        for nota in item.get("notas", []):
            NotaRecebimento.objects.create(
                item=item_obj,
                numero=str(nota.get("numero") or ""),
                quantidade=_decimal(nota.get("quantidade")),
                data_recebimento=_parse_dt(nota.get("dataRecebimento")),
                foto_path=nota.get("fotoPath") or "",
            )
    return contagem


def _token_valido(request):
    token = request.headers.get("X-App-Token", "")
    return bool(settings.APP_API_TOKEN) and token == settings.APP_API_TOKEN


def _unauthorized():
    return JsonResponse({"ok": False, "erro": "Nao autorizado"}, status=401)


def _parse_dt(value):
    if not value:
        return None
    parsed = parse_datetime(value)
    if parsed is None:
        raise ValueError(f"Data invalida: {value}")
    return parsed


def _decimal(value):
    if value in (None, ""):
        return Decimal("0")
    return Decimal(str(value))


def _sum_decimal(itens, campo):
    total = Decimal("0")
    for item in itens:
        total += _decimal(item.get(campo))
    return total


def _contagem_json(contagem):
    return {
        "id": contagem.id,
        "sessaoAppId": contagem.sessao_app_id,
        "operadorNome": contagem.operador_nome,
        "operadorMatricula": contagem.operador_matricula,
        "dataInicio": contagem.data_inicio.isoformat(),
        "dataFimReal": contagem.data_fim_real.isoformat() if contagem.data_fim_real else None,
        "status": contagem.status,
        "totalItens": contagem.total_itens,
        "totalContado": float(contagem.total_contado),
        "totalRecebido": float(contagem.total_recebido),
    }


def _contagem_publica_json(contagem):
    itens = list(contagem.itens.all())
    alertas = sum(1 for item in itens if item.status == "alerta")
    bloqueios = sum(1 for item in itens if item.status == "bloqueado")
    justificados = sum(1 for item in itens if item.status == "justificado")
    return {
        "id": contagem.id,
        "sessaoAppId": contagem.sessao_app_id,
        "titulo": f"Contagem {contagem.data_inicio:%d/%m/%Y} - {contagem.operador_nome}",
        "operador": contagem.operador_nome,
        "matricula": contagem.operador_matricula,
        "data": contagem.data_inicio.date().isoformat(),
        "dataInicio": contagem.data_inicio.isoformat(),
        "dataFimReal": contagem.data_fim_real.isoformat() if contagem.data_fim_real else None,
        "status": contagem.status,
        "totalMateriais": contagem.total_itens,
        "totalContado": float(contagem.total_contado),
        "totalRecebido": float(contagem.total_recebido),
        "alertas": alertas,
        "bloqueios": bloqueios,
        "justificados": justificados,
        "observacao": "Recebida automaticamente pelo app",
    }


def _item_json(item):
    return {
        "materialCodigo": item.material_codigo,
        "fornecedor": item.fornecedor,
        "materialDescricao": item.material_descricao,
        "estoqueAnterior": float(item.estoque_anterior),
        "estoqueContado": float(item.estoque_contado),
        "recebimentoTotal": float(item.recebimento_total),
        "notas": [_nota_json(nota) for nota in item.notas.all()],
        "consumoEstimado": float(item.consumo_estimado),
        "status": item.status,
        "justificativa": item.justificativa,
    }


def _nota_json(nota):
    return {
        "numero": nota.numero,
        "quantidade": float(nota.quantidade),
        "dataRecebimento": nota.data_recebimento.isoformat() if nota.data_recebimento else None,
        "fotoPath": nota.foto_path,
    }


def _gerar_excel_contagem(contagem):
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo"
    _append_rows(
        ws,
        [
            ["Campo", "Valor"],
            ["Sessao", contagem.sessao_app_id],
            ["Operador", contagem.operador_nome],
            ["Matricula", contagem.operador_matricula],
            ["Inicio", _excel_datetime(contagem.data_inicio)],
            ["Fim", _excel_datetime(contagem.data_fim_real)],
            ["Status", contagem.status],
            ["Total de itens", contagem.total_itens],
            ["Total contado", contagem.total_contado],
            ["Total recebido", contagem.total_recebido],
        ],
    )

    itens_ws = wb.create_sheet("Itens")
    _append_rows(
        itens_ws,
        [
            [
                "Fornecedor",
                "Codigo",
                "Material",
                "Estoque anterior",
                "Recebimento",
                "Soma NFs",
                "Estoque contado",
                "Consumo estimado",
                "Status",
                "Justificativa",
                "Observacao",
                "Horario",
                "Foto",
            ]
        ],
    )
    for item in contagem.itens.all():
        itens_ws.append(
            [
                item.fornecedor,
                item.material_codigo,
                item.material_descricao,
                item.estoque_anterior,
                item.recebimento_total,
                item.soma_nf,
                item.estoque_contado,
                item.consumo_estimado,
                item.status,
                item.justificativa,
                item.observacao,
                _excel_datetime(item.timestamp_item),
                item.foto_path,
            ]
        )

    notas_ws = wb.create_sheet("Notas fiscais")
    _append_rows(
        notas_ws,
        [["Fornecedor", "Codigo", "Material", "NF/GR", "Quantidade", "Data", "Foto"]],
    )
    for item in contagem.itens.all():
        for nota in item.notas.all():
            notas_ws.append(
                [
                    item.fornecedor,
                    item.material_codigo,
                    item.material_descricao,
                    nota.numero,
                    nota.quantidade,
                    _excel_datetime(nota.data_recebimento),
                    nota.foto_path,
                ]
            )

    for sheet in wb.worksheets:
        _format_sheet(sheet)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _append_rows(ws, rows):
    for row in rows:
        ws.append(row)


def _format_sheet(ws):
    header_fill = PatternFill("solid", fgColor="DCEFE9")
    header_font = Font(bold=True, color="0C3D31")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    ws.freeze_panes = "A2"
    for column_cells in ws.columns:
        width = 12
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(len(value) + 2, 42))
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width


def _excel_datetime(value):
    if value is None:
        return None
    return timezone.localtime(value).replace(tzinfo=None)
